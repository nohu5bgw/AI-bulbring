"""
MNP-branded Excel writeup exporter.
Accepts one or more StatementData objects (batch mode).
Produces a multi-sheet .xlsx:
  • One "Writeup" sheet per statement  (or combined if client_name matches)
  • One "Summary" sheet with cross-statement totals by category
"""

import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from typing import List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Brand colours ─────────────────────────────────────────────────────────────
MNP_RED     = "C8102E"
MNP_BLACK   = "1A1A1A"
MNP_GREY    = "F5F5F5"
MNP_ALT_ROW = "F9F9F9"
COL_GREEN   = "1E7B34"
COL_ORANGE  = "E65C00"
COL_WHITE   = "FFFFFF"
CURRENCY_FMT = '#,##0.00'

_THIN   = Side(style="thin", color="DDDDDD")
_BORDER = Border(bottom=Side(style="thin", color="EEEEEE"))


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=MNP_BLACK, size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _fmt_period(raw: str) -> str:
    """'NOV 18 TO DEC 15, 2025' → 'Nov 18 – Dec 15, 2025'"""
    raw = raw.strip()
    m = re.match(
        r'([A-Z]{3})\s+(\d{1,2})\s+(?:TO|THROUGH|[-–])\s*([A-Z]{3})\s+(\d{1,2}),?\s*(\d{4})?',
        raw, re.IGNORECASE
    )
    if m:
        y = m.group(5) or str(datetime.now().year)
        return (f"{m.group(1).capitalize()} {m.group(2)} – "
                f"{m.group(3).capitalize()} {m.group(4)}, {y}")
    return raw


# ── Public API ────────────────────────────────────────────────────────────────

def export_excel(statements, output_path: str) -> str:
    """
    statements : StatementData  OR  list[StatementData]
    output_path: destination .xlsx path
    Returns output_path on success.
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl not installed — run: pip install openpyxl")

    # Normalise to list
    if not isinstance(statements, list):
        statements = [statements]

    wb = Workbook()
    wb.remove(wb.active)  # strip default empty sheet

    # One writeup sheet per statement
    used_names = {}
    for idx, stmt in enumerate(statements):
        raw_name = stmt.source_file or f"Statement {idx+1}"
        # Strip extension, truncate to 28 chars (Excel sheet name limit is 31)
        sheet_label = Path(raw_name).stem[:28]
        # Deduplicate sheet names
        if sheet_label in used_names:
            used_names[sheet_label] += 1
            sheet_label = f"{sheet_label[:25]} ({used_names[sheet_label]})"
        else:
            used_names[sheet_label] = 1
        _build_writeup_sheet(wb, stmt, sheet_label)

    # Combined summary (always last)
    _build_summary_sheet(wb, statements)

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Sheet: Bank Writeup (one per statement)
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    ("Trans Date",    12),
    ("Post Date",     12),
    ("Cardholder",    26),
    ("Description",   38),
    ("Category",      22),
    ("For. Curr.",    10),
    ("For. Amt",      11),
    ("Exch. Rate",    11),
    ("Amount (CAD)",  14),
]
N_COLS   = len(COLUMNS)
LAST_COL = get_column_letter(N_COLS)


def _build_writeup_sheet(wb, data, sheet_title: str):
    ws = wb.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False

    period = _fmt_period(data.statement_period) if data.statement_period else "—"
    cards  = " / ".join(f"…{c[-4:]}" for c in data.card_numbers) if data.card_numbers else ""

    # ── Header block (rows 1-6) ───────────────────────────────────────────────
    def _merge(r, text, bold=False, size=11, fg=MNP_BLACK, bg=None, align="left"):
        ws.merge_cells(f"A{r}:{LAST_COL}{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font      = _font(bold=bold, size=size, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        if bg:
            c.fill = _fill(bg)
        ws.row_dimensions[r].height = 26 if size >= 16 else 18

    _merge(1, "MNP LLP", bold=True, size=20, fg=MNP_RED)
    _merge(2, "Bank Writeup — Credit Card Analysis", bold=True, size=13)
    _merge(3, f"Client:  {data.client_name}" if data.client_name else "Client:  —")
    _merge(4, f"Statement Period:  {period}")
    _merge(5, f"Card:  {cards}" if cards else "Card:  —")
    ws.row_dimensions[6].height = 8  # spacer

    # ── Column headers (row 7) ────────────────────────────────────────────────
    for ci, (label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=7, column=ci, value=label)
        c.font      = _font(bold=True, color=COL_WHITE, size=10)
        c.fill      = _fill(MNP_BLACK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    # ── Transactions ──────────────────────────────────────────────────────────
    cardholder_totals = defaultdict(float)
    total_purchases = total_payments = 0.0
    current_cardholder = None
    row = 8

    def _section_header(name, r):
        ws.merge_cells(f"A{r}:{LAST_COL}{r}")
        c = ws.cell(row=r, column=1, value=f"  {name.upper()}")
        c.font      = _font(bold=True, size=10, color=MNP_RED)
        c.fill      = _fill("FEF0F0")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 15
        return r + 1

    for idx, t in enumerate(data.transactions):
        if t.cardholder != current_cardholder:
            current_cardholder = t.cardholder
            row = _section_header(current_cardholder, row)

        bg       = MNP_ALT_ROW if idx % 2 else COL_WHITE
        is_cr    = t.amount_cad < 0
        is_unc   = t.category == "Uncategorized" and not is_cr
        ink      = COL_GREEN if is_cr else (COL_ORANGE if is_unc else MNP_BLACK)

        vals = [
            t.transaction_date, t.posting_date, t.cardholder, t.description,
            t.category,
            t.foreign_currency or "",
            t.foreign_amount   if t.foreign_amount else "",
            t.exchange_rate    if t.exchange_rate  else "",
            t.amount_cad,
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _font(color=ink, size=10)
            c.fill      = _fill(bg)
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if ci == N_COLS else "left")
            c.border    = _BORDER
            if ci == N_COLS:
                c.number_format = CURRENCY_FMT
            elif ci == 7 and val != "":
                c.number_format = CURRENCY_FMT
            elif ci == 8 and val != "":
                c.number_format = "0.000000"
        ws.row_dimensions[row].height = 15

        cardholder_totals[t.cardholder] += t.amount_cad
        if t.amount_cad >= 0:
            total_purchases += t.amount_cad
        else:
            total_payments  += t.amount_cad
        row += 1

    # ── Totals block ──────────────────────────────────────────────────────────
    row += 1

    def _total_row(label, amount, big=False):
        nonlocal row
        ws.merge_cells(f"A{row}:{get_column_letter(N_COLS - 1)}{row}")
        lc = ws.cell(row=row, column=1, value=label)
        ac = ws.cell(row=row, column=N_COLS, value=amount)
        sz = 11 if big else 10
        lc.font = _font(bold=True, size=sz, color=MNP_RED)
        lc.fill = _fill("FEF0F0")
        lc.alignment = Alignment(horizontal="right", vertical="center")
        ac.font = _font(bold=True, size=sz)
        ac.fill = _fill("FEF0F0")
        ac.number_format = CURRENCY_FMT
        ac.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    for ch, tot in sorted(cardholder_totals.items()):
        _total_row(f"  SUBTOTAL — {ch}:", tot)

    row += 1
    _total_row("TOTAL PURCHASES:", total_purchases)
    _total_row("TOTAL PAYMENTS:", total_payments)

    # Net row (red bar)
    ws.merge_cells(f"A{row}:{get_column_letter(N_COLS - 1)}{row}")
    lc = ws.cell(row=row, column=1, value="NET BALANCE:")
    ac = ws.cell(row=row, column=N_COLS, value=total_purchases + total_payments)
    for c in (lc, ac):
        c.font  = _font(bold=True, size=12, color=COL_WHITE)
        c.fill  = _fill(MNP_RED)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    ac.alignment = Alignment(horizontal="right", vertical="center")
    ac.number_format = CURRENCY_FMT
    ws.row_dimensions[row].height = 20

    # ── Page setup ────────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_area              = f"A1:{LAST_COL}{row}"
    ws.oddFooter.center.text  = "MNP LLP CONFIDENTIAL"
    ws.page_margins           = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet: Summary (combined across all statements)
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(wb, statements: List):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    row = 1

    def cell(r, col, value, bold=False, size=11, fg=MNP_BLACK, bg=None,
             fmt=None, align="left"):
        c = ws.cell(row=r, column=col, value=value)
        c.font      = _font(bold=bold, size=size, color=fg)
        c.alignment = Alignment(
            horizontal=align if col == 1 else "right", vertical="center"
        )
        if bg:
            c.fill = _fill(bg)
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 18
        return c

    def row_pair(label, amount, bold=False, lfg=MNP_BLACK, afg=MNP_BLACK, bg=None):
        nonlocal row
        cell(row, 1, label, bold=bold, fg=lfg, bg=bg)
        cell(row, 2, amount, bold=bold, fg=afg, bg=bg, fmt=CURRENCY_FMT)
        row += 1

    def blank(n=1):
        nonlocal row
        row += n

    def section_header(text):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font      = _font(bold=True, size=10, color=COL_WHITE)
        c.fill      = _fill(MNP_BLACK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value="MNP LLP — Combined Statement Summary")
    c.font      = _font(bold=True, size=16, color=MNP_RED)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    row = 2

    # Print date
    cell(row, 1, f"Generated: {datetime.now().strftime('%B %d, %Y')}",
         size=10, fg="888888")
    row += 1
    cell(row, 1, f"Statements included: {len(statements)}", size=10, fg="888888")
    row += 1
    blank()

    # ── Per-statement summary ─────────────────────────────────────────────────
    section_header("STATEMENT BREAKDOWN")
    for stmt in statements:
        period = _fmt_period(stmt.statement_period) if stmt.statement_period else "—"
        client = stmt.client_name or "Unknown"
        source = stmt.source_file or "—"
        n      = len(stmt.transactions)
        net    = sum(t.amount_cad for t in stmt.transactions)

        cell(row, 1, f"{source}", bold=True, size=10)
        cell(row, 2, net, bold=True, size=10,
             fg=COL_GREEN if net < 0 else MNP_BLACK, fmt=CURRENCY_FMT)
        row += 1

        cell(row, 1, f"  Client: {client}  |  Period: {period}  |  {n} transactions",
             size=9, fg="666666")
        row += 1

    blank()

    # ── Aggregate balances ────────────────────────────────────────────────────
    section_header("AGGREGATE TOTALS")
    total_prev    = sum(s.previous_balance  for s in statements)
    total_pay     = sum(s.payments_credits  for s in statements)
    total_purch   = sum(s.purchases_debits  for s in statements)
    total_new     = sum(s.new_balance       for s in statements)
    parsed_purch  = sum(t.amount_cad for s in statements for t in s.transactions if t.amount_cad > 0)
    parsed_pay    = sum(t.amount_cad for s in statements for t in s.transactions if t.amount_cad < 0)

    if total_prev:
        row_pair("Previous Balance (total):", total_prev)
    row_pair("Total Payments & Credits:", parsed_pay,   afg=COL_GREEN)
    row_pair("Total Purchases & Debits:", parsed_purch)
    if total_new:
        row_pair("Total New Balance:", total_new, bold=True, lfg=MNP_RED, afg=MNP_RED, bg="FEF0F0")
    blank()

    # ── Category breakdown ────────────────────────────────────────────────────
    section_header("BREAKDOWN BY CATEGORY (ALL STATEMENTS)")
    by_cat: dict = defaultdict(float)
    for stmt in statements:
        for t in stmt.transactions:
            by_cat[t.category] += t.amount_cad

    def _sort_key(kv):
        k = kv[0]
        if k == "Payment / Credit": return (2, k)
        if k == "Uncategorized":    return (1, k)
        return (0, k)

    for cat, total in sorted(by_cat.items(), key=_sort_key):
        is_unc = cat == "Uncategorized"
        is_pay = cat == "Payment / Credit"
        lfg    = COL_ORANGE if is_unc else (MNP_BLACK)
        afg    = COL_ORANGE if is_unc else (COL_GREEN if is_pay else MNP_BLACK)
        bg     = "FFFCE8" if is_unc else None
        label  = f"{cat}:  ◄ review" if is_unc else f"{cat}:"
        row_pair(label, total, lfg=lfg, afg=afg, bg=bg)

    blank()
    cell(row, 1, "Prepared using MNP Bank Writeup Tool", size=9, fg="AAAAAA")
    row += 1

    ws.page_setup.orientation = "portrait"
    ws.print_area              = f"A1:B{row}"
    ws.oddFooter.center.text  = "MNP LLP CONFIDENTIAL"
    ws.page_margins           = PageMargins(left=0.75, right=0.75, top=0.75, bottom=0.75)
